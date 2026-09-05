"""Fill the official two-track workbook without inventing human review."""
from __future__ import annotations

import math
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from mva_track1.common import Track1Error


def write_methods(template, output, answers: dict) -> dict:
    book = load_workbook(template)
    filled = 0
    for sheet in book:
        expected = range(7, 20 if sheet.title == 'Track 1 methods' else 18)
        supplied = answers.get(sheet.title, {})
        if set(supplied) != set(expected):
            raise Track1Error('Methods workbook answer mapping does not match the official questions')
        for row in expected:
            value = supplied[row].strip()
            if not value:
                raise Track1Error('A methods answer is empty')
            if row == max(expected) and len(value.split()) > 500:
                raise Track1Error('Methods abstract exceeds the official 500-word limit')
            cell = sheet.cell(row, 2)
            cell.value = value
            cell.data_type = 's'  # A model/user answer must never become an Excel formula.
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            a_width = float(sheet.column_dimensions['A'].width or 40)
            b_width = float(sheet.column_dimensions['B'].width or 70)
            lines = max(math.ceil(len(str(sheet.cell(row, 1).value or '')) / max(20, a_width)),
                        math.ceil(len(value) / max(20, b_width))) + value.count('\n') + 2
            sheet.row_dimensions[row].height = min(409, max(45, lines * 15))
            filled += 1
    book.save(output)
    checked = load_workbook(output)
    for sheet in checked:
        for row, expected in answers[sheet.title].items():
            if sheet.cell(row, 2).value != expected.strip() or sheet.cell(row, 2).data_type == 'f':
                raise Track1Error('Workbook round-trip validation failed')
    return {'sheets': checked.sheetnames, 'answers_filled': filled, 'round_trip_verified': True}
